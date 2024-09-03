import openpyxl
import datetime


def generate_data():
    wb = openpyxl.load_workbook(
        'file.xlsx'
    )
    sheet = wb.active
    data = []

    if is_even_week():
        if datetime.date.today().weekday() == 0:
            for row in sheet.iter_rows(min_row=2, max_row=6, min_col=5, max_col=6):
                for cell in range(0, 2, len(row)):
                    if row[cell + 1].value:
                        t = [row[cell].value, row[cell + 1].value]
                        data.append(t)
        elif datetime.date.today().weekday() == 1:
            for row in sheet.iter_rows(min_row=7, max_row=11, min_col=5, max_col=6):
                for cell in range(0, 2, len(row)):
                    if row[cell + 1].value:
                        t = [row[cell].value, row[cell + 1].value]
                        data.append(t)
        elif datetime.date.today().weekday() == 2:
            data.append(["8:10 - 18:40", "Военная подготовка"])
        elif datetime.date.today().weekday() == 3:
            for row in sheet.iter_rows(min_row=2, max_row=6, min_col=10, max_col=11):
                for cell in range(0, 2, len(row)):
                    if row[cell + 1].value:
                        t = [row[cell].value, row[cell + 1].value]
                        data.append(t)
        elif datetime.date.today().weekday() == 4:
            for row in sheet.iter_rows(min_row=7, max_row=11, min_col=10, max_col=11):
                for cell in range(0, 2, len(row)):
                    if row[cell + 1].value:
                        t = [row[cell].value, row[cell + 1].value]
                        data.append(t)
        elif datetime.date.today().weekday() == 5:
            for row in sheet.iter_rows(min_row=12,
                                       max_row=14,
                                       min_col=10,
                                       max_col=11):
                for cell in range(0, 2, len(row)):
                    if row[cell + 1].value:
                        t = [row[cell].value, row[cell + 1].value]
                        data.append(t)
    else:
        if datetime.date.today().weekday() == 0:
            for row in sheet.iter_rows(min_row=2, max_row=5, min_col=5, max_col=7):
                for cell in range(0, 3, len(row)):
                    if row[cell + 2].value:
                        t = [row[cell].value, row[cell + 2].value]
                        data.append(t)
        elif datetime.date.today().weekday() == 1:
            for row in sheet.iter_rows(min_row=6, max_row=9, min_col=5, max_col=7):
                for cell in range(0, 3, len(row)):
                    if row[cell + 2].value:
                        t = [row[cell].value, row[cell + 2].value]
                        data.append(t)
        elif datetime.date.today().weekday() == 2:
            for row in sheet.iter_rows(min_row=10, max_row=13, min_col=5, max_col=7):
                for cell in range(0, 3, len(row)):
                    if row[cell + 2].value:
                        t = [row[cell].value, row[cell + 2].value]
                        data.append(t)
        elif datetime.date.today().weekday() == 3:
            data.append(["8:10 - 18:40", "Военная подготовка"])
        elif datetime.date.today().weekday() == 4:
            for row in sheet.iter_rows(min_row=6, max_row=9, min_col=10, max_col=12):
                for cell in range(0, 3, len(row)):
                    if row[cell + 2].value:
                        t = [row[cell].value, row[cell + 2].value]
                        data.append(t)
        elif datetime.date.today().weekday() == 5:
            for row in sheet.iter_rows(min_row=10,
                                       max_row=13,
                                       min_col=10,
                                       max_col=12):
                for cell in range(0, 3, len(row)):
                    if row[cell + 2].value:
                        t = [row[cell].value, row[cell + 2].value]
                        data.append(t)

    print(data)
    return data


def is_even_week():
    today = datetime.date.today()
    iso_calendar = today.isocalendar()
    week_number = iso_calendar[1]

    return week_number % 2 == 0
